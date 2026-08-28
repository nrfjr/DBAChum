import io

import pytest
from openpyxl import Workbook, load_workbook
from fastapi import UploadFile

from app.core.exceptions import AppError
from app.schemas.provisioning import BulkProvisionRequest, BulkProvisionRowInput
from app.services import bulk_provisioning


@pytest.mark.asyncio
async def test_csv_import_requires_identity_headers(monkeypatch):
    upload = UploadFile(
        filename="users.csv",
        file=io.BytesIO(b"first_name,last_name\nJuan,Santos\n"),
    )

    with pytest.raises(AppError) as exc:
        await bulk_provisioning.import_bulk_provision_file(None, "db1", upload)

    assert exc.value.code == "BULK_IMPORT_HEADERS_MISSING"


@pytest.mark.asyncio
async def test_csv_import_generates_username_and_password(monkeypatch):
    async def fake_target(database, connection_id):
        return {"id": connection_id}

    async def fake_existing(connection, usernames):
        return set()

    monkeypatch.setattr(bulk_provisioning, "get_oracle_target", fake_target)
    monkeypatch.setattr(bulk_provisioning, "find_existing_oracle_users", fake_existing)

    upload = UploadFile(
        filename="users.csv",
        file=io.BytesIO(
            b"employee_id,first_name,middle_name,last_name,reference_user,password\n"
            b"12345,Juan,M,Santos,,\n"
        ),
    )

    result = await bulk_provisioning.import_bulk_provision_file(None, "db1", upload)

    assert result.row_count == 1
    assert result.valid_count == 1
    row = result.rows[0]
    assert row.username == "JMSANTOS12345"
    assert row.reference_user is None
    assert row.password_mode == "generated"
    assert len(row.password) == 8


@pytest.mark.asyncio
async def test_import_marks_existing_oracle_username_invalid(monkeypatch):
    async def fake_target(database, connection_id):
        return {"id": connection_id}

    async def fake_existing(connection, usernames):
        return set(usernames)

    monkeypatch.setattr(bulk_provisioning, "get_oracle_target", fake_target)
    monkeypatch.setattr(bulk_provisioning, "find_existing_oracle_users", fake_existing)

    upload = UploadFile(
        filename="users.csv",
        file=io.BytesIO(b"employee_id,first_name,last_name\n12345,Juan,Santos\n"),
    )

    result = await bulk_provisioning.import_bulk_provision_file(None, "db1", upload)

    assert result.invalid_count == 1
    assert result.rows[0].errors["username"] == "This Oracle username already exists."


def test_bulk_common_reference_is_required_when_enabled():
    with pytest.raises(ValueError):
        BulkProvisionRequest(
            use_common_reference=True,
            rows=[
                BulkProvisionRowInput(
                    row_number=2,
                    employee_id="12345",
                    first_name="Juan",
                    last_name="Santos",
                    password="abc12345",
                )
            ],
        )


@pytest.mark.asyncio
async def test_xlsx_import_preserves_text_employee_id_and_zero_number_format(monkeypatch):
    async def fake_target(database, connection_id):
        return {"id": connection_id}

    async def fake_existing(connection, usernames):
        return set()

    monkeypatch.setattr(bulk_provisioning, "get_oracle_target", fake_target)
    monkeypatch.setattr(bulk_provisioning, "find_existing_oracle_users", fake_existing)

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["employee_id", "first_name", "last_name"])
    sheet.append([289, "John-Doe", "Last-name"])
    sheet["A2"].number_format = "0000"
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)

    upload = UploadFile(filename="users.xlsx", file=buffer)
    result = await bulk_provisioning.import_bulk_provision_file(None, "db1", upload)

    assert result.rows[0].employee_id == "0289"
    assert result.rows[0].username == "JLASTNAME0289"
    assert result.rows[0].valid is True


def test_bulk_template_xlsx_formats_employee_id_as_text():
    content = bulk_provisioning.build_bulk_template_xlsx()
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    assert sheet["A2"].value == "001234"
    assert sheet["A2"].number_format == "@"
    workbook.close()
