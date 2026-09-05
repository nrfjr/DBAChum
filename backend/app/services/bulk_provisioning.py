from __future__ import annotations

import csv
import io
import secrets
import string
from typing import Iterable

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.connectors.oracle_provisioning import (
    find_existing_oracle_users,
    get_oracle_reference_user,
    is_sensitive_reference_role,
    normalize_oracle_identifier,
)
from app.core.exceptions import AppError
from app.schemas.oracle_dba import OracleCreateUserRequest
from app.schemas.provisioning import (
    BulkProvisionExecutionResponse,
    BulkProvisionExecutionRow,
    BulkProvisionImportResponse,
    BulkProvisionImportRow,
    BulkProvisionPreviewResponse,
    BulkProvisionPreviewRow,
    BulkProvisionRequest,
    ProvisioningExecuteRequest,
    ProvisioningPreviewRequest,
)
from app.schemas.user import UserResponse
from app.services.database_connections import get_database_connection
from app.services.ldap_ldif import normalize_employee_id, normalize_person_name
from app.services.oracle_dba import get_oracle_target, provision_oracle_user
from app.services.provisioning import get_provisioning_profile
from app.services.provisioning_execution import execute_provisioning_profile
from app.services.provisioning_preview import (
    build_provisioning_preview,
    generate_provisioning_username,
)


REQUIRED_HEADERS = ["employee_id", "first_name", "last_name"]
OPTIONAL_HEADERS = ["middle_name", "password", "reference_user"]
MAX_BULK_ROWS = 500
MAX_UPLOAD_BYTES = 5 * 1024 * 1024

HEADER_ALIASES = {
    "employeeid": "employee_id",
    "id": "employee_id",
    "firstname": "first_name",
    "givenname": "first_name",
    "middlename": "middle_name",
    "middleinitial": "middle_name",
    "lastname": "last_name",
    "surname": "last_name",
    "password": "password",
    "referenceuser": "reference_user",
    "refuser": "reference_user",
}


def _normalize_header(value: object) -> str:
    compact = "".join(character.lower() for character in str(value or "") if character.isalnum())
    return HEADER_ALIASES.get(compact, compact)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _xlsx_employee_id_text(cell) -> str:

    value = cell.value
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        number_format = str(getattr(cell, "number_format", "") or "").strip()
        if number_format and set(number_format) <= {"0"} and "." not in number_format:
            width = len(number_format)
            try:
                return str(int(value)).zfill(width)
            except (TypeError, ValueError):
                pass
    return _cell_text(value)


def _generate_password() -> str:
    letters = string.ascii_lowercase
    digits = string.digits
    return "".join(secrets.choice(letters) for _ in range(3)) + "".join(
        secrets.choice(digits) for _ in range(5)
    )


def _validate_name(raw: str, field_label: str, *, required: bool) -> tuple[str, str | None]:

    cleaned = normalize_person_name(raw)
    if required and not cleaned:
        return "", f"{field_label} is required."
    if raw and any(character.isdigit() for character in raw):
        return cleaned or "", f"{field_label} cannot contain numbers."
    if raw and not cleaned:
        return "", f"{field_label} must contain at least one letter."
    return cleaned or "", None


def _validate_employee_id(raw: str) -> tuple[str, str | None]:
    cleaned = normalize_employee_id(raw)
    if not cleaned:
        return "", "Employee ID is required."
    if raw.strip() != cleaned:
        return cleaned, "Employee ID must contain letters and numbers only."
    return cleaned, None


def _validate_password(raw: str) -> str | None:
    if len(raw) < 8:
        return "Password must contain at least 8 characters."
    if len(raw) > 128:
        return "Password cannot exceed 128 characters."
    if '"' in raw or any(ord(character) < 32 for character in raw):
        return "Password cannot contain double quotes or control characters."
    return None


def _rows_from_csv(contents: bytes) -> tuple[list[str], list[list[object]]]:
    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise AppError(
            "CSV files must use UTF-8 encoding.",
            code="BULK_IMPORT_ENCODING_INVALID",
            status_code=400,
        ) from exc
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return [str(item or "") for item in rows[0]], rows[1:]


def _rows_from_xlsx(contents: bytes) -> tuple[list[str], list[list[object]]]:
    try:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise AppError(
            "The XLSX file could not be read. Upload a valid Excel workbook.",
            code="BULK_IMPORT_XLSX_INVALID",
            status_code=400,
        ) from exc
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=False)
        header_cells = next(iterator, None)
        if header_cells is None:
            return [], []
        headers = [str(cell.value or "") for cell in header_cells]
        employee_id_index = next(
            (index for index, header in enumerate(headers) if _normalize_header(header) == "employee_id"),
            None,
        )
        rows: list[list[object]] = []
        for cells in iterator:
            converted: list[object] = []
            for index, cell in enumerate(cells):
                if employee_id_index is not None and index == employee_id_index:
                    converted.append(_xlsx_employee_id_text(cell))
                else:
                    converted.append(_cell_text(cell.value))
            rows.append(converted)
        return headers, rows
    finally:
        workbook.close()


def _map_headers(headers: Iterable[object]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    duplicates: set[str] = set()
    for index, raw in enumerate(headers):
        canonical = _normalize_header(raw)
        if not canonical:
            continue
        if canonical in mapped:
            duplicates.add(canonical)
        mapped[canonical] = index
    if duplicates:
        names = ", ".join(sorted(duplicates))
        raise AppError(
            f"Duplicate spreadsheet header(s): {names}.",
            code="BULK_IMPORT_DUPLICATE_HEADERS",
            status_code=400,
        )
    missing = [header for header in REQUIRED_HEADERS if header not in mapped]
    if missing:
        raise AppError(
            "Missing required spreadsheet header(s): " + ", ".join(missing) + ".",
            code="BULK_IMPORT_HEADERS_MISSING",
            status_code=400,
        )
    return mapped


def _value(row: list[object], mapped: dict[str, int], key: str) -> str:
    index = mapped.get(key)
    if index is None or index >= len(row):
        return ""
    return _cell_text(row[index])


async def import_bulk_provision_file(
    database,
    connection_id: str,
    upload: UploadFile,
) -> BulkProvisionImportResponse:
    filename = upload.filename or "upload"
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in {"csv", "xlsx"}:
        raise AppError(
            "Bulk provisioning supports .xlsx and .csv files.",
            code="BULK_IMPORT_FILE_TYPE_INVALID",
            status_code=400,
        )

    contents = await upload.read(MAX_UPLOAD_BYTES + 1)
    if len(contents) > MAX_UPLOAD_BYTES:
        raise AppError(
            "Bulk provisioning files cannot exceed 5 MB.",
            code="BULK_IMPORT_FILE_TOO_LARGE",
            status_code=400,
        )

    headers, raw_rows = _rows_from_csv(contents) if extension == "csv" else _rows_from_xlsx(contents)
    mapped = _map_headers(headers)

    parsed: list[BulkProvisionImportRow] = []
    seen_usernames: dict[str, int] = {}
    for sheet_row_number, raw_row in enumerate(raw_rows, start=2):
        if not any(_cell_text(value) for value in raw_row):
            continue
        if len(parsed) >= MAX_BULK_ROWS:
            raise AppError(
                f"Bulk provisioning supports at most {MAX_BULK_ROWS} data rows per file.",
                code="BULK_IMPORT_TOO_MANY_ROWS",
                status_code=400,
            )

        errors: dict[str, str] = {}
        employee_id, error = _validate_employee_id(_value(raw_row, mapped, "employee_id"))
        if error:
            errors["employee_id"] = error
        first_name, error = _validate_name(_value(raw_row, mapped, "first_name"), "First name", required=True)
        if error:
            errors["first_name"] = error
        middle_name, error = _validate_name(_value(raw_row, mapped, "middle_name"), "Middle name", required=False)
        if error:
            errors["middle_name"] = error
        last_name, error = _validate_name(_value(raw_row, mapped, "last_name"), "Last name", required=True)
        if error:
            errors["last_name"] = error

        reference_user = _value(raw_row, mapped, "reference_user").upper() or None
        if reference_user:
            try:
                reference_user = normalize_oracle_identifier(reference_user, field_name="Reference user")
            except AppError as exc:
                errors["reference_user"] = exc.message

        supplied_password = _value(raw_row, mapped, "password")
        password_mode = "provided" if supplied_password else "generated"
        password = supplied_password or _generate_password()
        password_error = _validate_password(password)
        if password_error:
            errors["password"] = password_error

        username = None
        if not any(key in errors for key in ("employee_id", "first_name", "middle_name", "last_name")):
            try:
                username = generate_provisioning_username(
                    first_name=first_name,
                    middle_name=middle_name or None,
                    last_name=last_name,
                    employee_id=employee_id,
                )
            except AppError as exc:
                errors["username"] = exc.message

        if username:
            previous = seen_usernames.get(username)
            if previous is not None:
                errors["username"] = f"Generated username duplicates spreadsheet row {previous}."
            else:
                seen_usernames[username] = sheet_row_number

        parsed.append(BulkProvisionImportRow(
            row_number=sheet_row_number,
            employee_id=employee_id,
            first_name=first_name,
            middle_name=middle_name or None,
            last_name=last_name,
            reference_user=reference_user,
            password=password,
            password_mode=password_mode,
            username=username,
            valid=not errors,
            errors=errors,
        ))

    if not parsed:
        raise AppError(
            "The spreadsheet does not contain any data rows.",
            code="BULK_IMPORT_EMPTY",
            status_code=400,
        )

    target = await get_oracle_target(database, connection_id)
    candidates = [row.username for row in parsed if row.username and row.valid]
    existing = await find_existing_oracle_users(target, [value for value in candidates if value])
    for row in parsed:
        if row.username and row.username in existing:
            row.errors["username"] = "This Oracle username already exists."
            row.valid = False

    valid_count = sum(1 for row in parsed if row.valid)
    return BulkProvisionImportResponse(
        filename=filename,
        required_headers=REQUIRED_HEADERS,
        optional_headers=OPTIONAL_HEADERS,
        row_count=len(parsed),
        valid_count=valid_count,
        invalid_count=len(parsed) - valid_count,
        rows=parsed,
    )



def build_bulk_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Users"
    headers = ["employee_id", "first_name", "middle_name", "last_name", "password", "reference_user"]
    sheet.append(headers)
    sheet.append(["001234", "Juan", "M", "Santos", "", ""])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for cell in sheet["A"]:
        cell.number_format = "@"
    widths = {"A": 18, "B": 20, "C": 20, "D": 24, "E": 20, "F": 22}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def build_bulk_results_xlsx(rows: list[dict[str, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Provisioning results"
    headers = [
        "row", "employee_id", "first_name", "middle_name", "last_name",
        "username", "initial_password", "status", "run_or_audit", "error",
    ]
    sheet.append(headers)
    for item in rows:
        sheet.append([item.get(header, "") for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for cell in sheet["B"]:
        cell.number_format = "@"
    for cell in sheet["G"]:
        cell.number_format = "@"
    widths = [8, 18, 20, 20, 24, 32, 20, 14, 36, 50]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()

def _effective_reference(data: BulkProvisionRequest, row) -> str | None:
    value = data.common_reference_user if data.use_common_reference else row.reference_user
    if not value:
        return None
    return normalize_oracle_identifier(value, field_name="Reference user")


async def preview_bulk_provisioning(
    database,
    connection_id: str,
    data: BulkProvisionRequest,
    operator: UserResponse,
    requester_ip: str | None = None,
) -> BulkProvisionPreviewResponse:
    target = await get_oracle_target(database, connection_id)
    profile_name = None
    if data.profile_id:
        profile = await get_provisioning_profile(database, data.profile_id)
        if profile.get("schema_connection_id") != connection_id:
            raise AppError(
                "The selected provisioning profile does not belong to this Oracle database.",
                code="PROVISIONING_PROFILE_WRONG_DATABASE",
                status_code=400,
            )
        profile_name = profile.get("name")

    generated: list[str] = []
    row_usernames: dict[int, str] = {}
    for row in data.rows:
        try:
            username = generate_provisioning_username(
                first_name=row.first_name,
                middle_name=row.middle_name,
                last_name=row.last_name,
                employee_id=row.employee_id,
            )
        except AppError:
            continue
        generated.append(username)
        row_usernames[row.row_number] = username
    existing = await find_existing_oracle_users(target, generated)

    seen: dict[str, int] = {}
    rows: list[BulkProvisionPreviewRow] = []
    for row in data.rows:
        errors: dict[str, str] = {}
        username = row_usernames.get(row.row_number)
        if not username:
            errors["username"] = "Unable to generate a valid Oracle username from this row."
        elif username in existing:
            errors["username"] = "This Oracle username already exists."
        elif username in seen:
            errors["username"] = f"Generated username duplicates batch row {seen[username]}."
        else:
            seen[username] = row.row_number

        reference_user = None
        try:
            reference_user = _effective_reference(data, row)
        except AppError as exc:
            errors["reference_user"] = exc.message

        roles: list[str] = []
        provisioning = None
        if not errors:
            try:
                if data.profile_id:
                    provisioning = await build_provisioning_preview(
                        database,
                        data.profile_id,
                        ProvisioningPreviewRequest(
                            username=username,
                            password=row.password,
                            first_name=row.first_name,
                            middle_name=row.middle_name,
                            last_name=row.last_name,
                            employee_id=row.employee_id,
                            reference_user=reference_user,
                            requestor=data.requestor,
                            request_reference=data.request_reference,
                            remarks=data.remarks,
                        ),
                        operator,
                        requester_ip=requester_ip,
                        parent_connection_id=connection_id,
                    )
                    if provisioning.account_exists:
                        errors["username"] = "This Oracle username already exists."
                    roles = [role.name for role in provisioning.roles if role.will_copy]
                elif reference_user:
                    reference = await get_oracle_reference_user(target, reference_user)
                    roles = [
                        role["name"] for role in reference["roles"]
                        if not is_sensitive_reference_role(role["name"])
                    ]
            except AppError as exc:
                key = "reference_user" if "REFERENCE" in exc.code else "row"
                errors[key] = exc.message

        rows.append(BulkProvisionPreviewRow(
            row_number=row.row_number,
            employee_id=row.employee_id,
            first_name=row.first_name,
            middle_name=row.middle_name,
            last_name=row.last_name,
            username=username,
            reference_user=reference_user,
            password_mode=row.password_mode,
            valid=not errors,
            errors=errors,
            roles=roles,
            provisioning=provisioning,
        ))

    valid_count = sum(1 for row in rows if row.valid)
    return BulkProvisionPreviewResponse(
        ready_to_execute=valid_count == len(rows) and bool(rows),
        row_count=len(rows),
        valid_count=valid_count,
        invalid_count=len(rows) - valid_count,
        profile_id=data.profile_id,
        profile_name=profile_name,
        rows=rows,
    )


async def execute_bulk_provisioning(
    database,
    connection_id: str,
    data: BulkProvisionRequest,
    operator: UserResponse,
    requester_ip: str | None = None,
) -> BulkProvisionExecutionResponse:
    live_preview = await preview_bulk_provisioning(
        database, connection_id, data, operator, requester_ip=requester_ip
    )
    if not live_preview.ready_to_execute:
        raise AppError(
            "Bulk provisioning is blocked because one or more rows failed the live preview.",
            code="BULK_PROVISIONING_PREVIEW_BLOCKED",
            status_code=409,
        )

    preview_by_row = {row.row_number: row for row in live_preview.rows}
    results: list[BulkProvisionExecutionRow] = []
    for row in data.rows:
        preview = preview_by_row[row.row_number]
        try:
            if data.profile_id:
                p = preview.provisioning
                assert p is not None
                result = await execute_provisioning_profile(
                    database,
                    connection_id,
                    data.profile_id,
                    ProvisioningExecuteRequest(
                        username=preview.username,
                        password=row.password,
                        first_name=row.first_name,
                        middle_name=row.middle_name,
                        last_name=row.last_name,
                        employee_id=row.employee_id,
                        reference_user=preview.reference_user,
                        requestor=data.requestor,
                        request_reference=data.request_reference,
                        remarks=data.remarks,
                        roles=preview.roles,
                        default_tablespace=p.default_tablespace,
                        temporary_tablespace=p.temporary_tablespace,
                        oracle_profile=p.oracle_profile,
                    ),
                    operator,
                    requester_ip=requester_ip,
                )
                results.append(BulkProvisionExecutionRow(
                    row_number=row.row_number,
                    username=result.username,
                    status=result.status,
                    run_id=result.run_id,
                    audit_id=result.audit_id,
                    error=result.error,
                ))
            else:
                created = await provision_oracle_user(
                    database,
                    connection_id,
                    OracleCreateUserRequest(
                        username=preview.username or "",
                        password=row.password,
                        reference_username=preview.reference_user,
                        roles=preview.roles,
                        request_reference=data.request_reference,
                        requestor_name=data.requestor,
                        remarks=data.remarks,
                        first_name=row.first_name,
                        middle_name=row.middle_name,
                        last_name=row.last_name,
                        employee_id=row.employee_id,
                        generate_ldif=False,
                        ldap_profile_id=None,
                    ),
                    operator,
                    requester_ip=requester_ip,
                )
                results.append(BulkProvisionExecutionRow(
                    row_number=row.row_number,
                    username=created["username"],
                    status="succeeded",
                    audit_id=created["audit_id"],
                ))
        except AppError as exc:

            results.append(BulkProvisionExecutionRow(
                row_number=row.row_number,
                username=preview.username,
                status="failed",
                error=exc.message,
            ))
        except Exception:
            results.append(BulkProvisionExecutionRow(
                row_number=row.row_number,
                username=preview.username,
                status="failed",
                error="Bulk provisioning failed unexpectedly for this row.",
            ))

    succeeded = sum(1 for row in results if row.status == "succeeded")
    partial = sum(1 for row in results if row.status == "partial")
    failed = sum(1 for row in results if row.status == "failed")
    overall = "succeeded" if succeeded == len(results) else "failed" if failed == len(results) else "partial"
    return BulkProvisionExecutionResponse(
        status=overall,
        row_count=len(results),
        succeeded_count=succeeded,
        partial_count=partial,
        failed_count=failed,
        rows=results,
    )
